"""
Daraz Campaign Manager blueprint for tb_track.

Workflow:
  1. Seller downloads the campaign Excel template from Daraz Seller Center.
  2. Uploads it here -> parsed, original xlsx kept on disk untouched.
  3. UI shows a friendly grid: English product name + image + color, current
     campaign price, recommended-price range, min-price (cost) floor, stock.
  4. Seller edits campaign prices (with validation + bulk tools).
  5. Downloads -> we mutate only the 'Campaign Price' column in the original
     workbook and stream it back. Daraz accepts it because every other cell is
     byte-for-byte preserved.
"""
from __future__ import annotations

import json
import os
import re
import time
import uuid
import zipfile
from datetime import datetime
from html import unescape
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import requests
from flask import Blueprint, abort, jsonify, render_template, request, send_file
from openpyxl import load_workbook

BASE_DIR = Path(os.getenv("CAMPAIGN_DIR", "campaign_uploads")).resolve()


def init_campaign_dirs() -> None:
    BASE_DIR.mkdir(parents=True, exist_ok=True)


LEGACY_COLUMN_MAP = {
    "seller_sku": 0,
    "sku_id": 1,
    "shop_sku": 2,
    "product_name": 3,
    "product_url": 4,
    "sales_price": 5,
    "campaign_price": 6,
    "recommended_price": 9,
    "stock": 10,
    "is_hero": 11,
    "category": 16,
}

HEADER_ALIASES = {
    "seller_sku": {"seller sku", "seller_sku"},
    "sku_id": {"sku id", "skuid", "sku_id"},
    "shop_sku": {"shop sku", "shop_sku"},
    "product_name": {"product name", "product_name"},
    "product_url": {"product url", "product_url", "product link", "productlink"},
    "sales_price": {"sales price", "sales_price", "current price", "price"},
    "campaign_price": {"campaign price", "campaign_price"},
    "recommended_price": {"recommended price", "recommended_price", "traffic-boosting price", "traffic boosting price"},
    "stock": {"stock", "current stock", "current_stock"},
    "is_hero": {"is hero product", "is hero", "is_hero_product", "hero product"},
    "category": {"category"},
}

HEADER_SCAN_LIMIT = 6
RECOMMENDED_PATTERN = re.compile(r"<\s*=?\s*([\d.]+)")


def parse_recommended_max(cell_value: Any) -> float | None:
    if cell_value is None:
        return None
    match = RECOMMENDED_PATTERN.search(str(cell_value))
    return float(match.group(1)) if match else None


_daraz_cache: dict[str, dict[str, Any]] = {}
ENRICHMENT_VERSION = 4


def _ensure_campaign_enrichment_version(campaign_id: str, meta: dict[str, Any]) -> dict[str, Any]:
    if meta.get("enrichment_version") == ENRICHMENT_VERSION:
        return meta

    for row in meta.get("rows", []):
        row["enrichment"] = None

    meta["enrichment_version"] = ENRICHMENT_VERSION
    _save_meta(campaign_id, meta)
    return meta


def _extract_meta_content(html: str, property_name: str) -> str | None:
    patterns = [
        rf'<meta[^>]+property=["\']{re.escape(property_name)}["\'][^>]+content=["\']([^"\']+)["\']',
        rf'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']{re.escape(property_name)}["\']',
        rf'<meta[^>]+name=["\']{re.escape(property_name)}["\'][^>]+content=["\']([^"\']+)["\']',
        rf'<meta[^>]+content=["\']([^"\']+)["\'][^>]+name=["\']{re.escape(property_name)}["\']',
    ]
    for pattern in patterns:
        match = re.search(pattern, html, flags=re.IGNORECASE)
        if match:
            return unescape(match.group(1)).strip()
    return None


def _extract_title_from_html(html: str) -> str | None:
    title = _extract_meta_content(html, "og:title")
    if title:
        return title
    match = re.search(r"<title>(.*?)</title>", html, flags=re.IGNORECASE | re.DOTALL)
    if match:
        return unescape(match.group(1)).strip()
    return None


def _extract_tracking_field_from_html(html: str, field_name: str) -> str | None:
    pattern = rf'"{re.escape(field_name)}":"([^"]+)"'
    match = re.search(pattern, html)
    if match:
        return unescape(match.group(1)).replace("\\/", "/").strip()
    return None


def _extract_variant_image_from_html(html: str, sku_id: str) -> str | None:
    if not sku_id:
        return None
    pattern = (
        rf'"{re.escape(str(sku_id))}":\{{'
        rf'"categoryId":"[^"]*",'
        rf'"image":"([^"]+)"'
    )
    match = re.search(pattern, html)
    if match:
        return unescape(match.group(1)).replace("\\/", "/").strip()
    return None


def enrich_from_daraz_row(row: dict[str, Any]) -> dict[str, Any]:
    seller_sku = str(row.get("seller_sku") or "").strip()
    sku_id = str(row.get("sku_id") or "").strip()
    product_url = str(row.get("product_url") or "").strip()
    cache_key = f"{product_url}|{sku_id}" if product_url else (seller_sku or sku_id)
    if not cache_key:
        return {}
    if cache_key in _daraz_cache:
        return _daraz_cache[cache_key]

    fallback = {
        "title": row.get("daraz_name") or seller_sku or "Untitled product",
        "image": None,
        "color": None,
        "variant_title": None,
    }

    if not product_url:
        _daraz_cache[cache_key] = fallback
        return fallback

    try:
        response = requests.get(
            product_url,
            timeout=12,
            headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
            },
        )
        response.raise_for_status()
        html = response.text
        title = (
            _extract_tracking_field_from_html(html, "pdt_name")
            or _extract_title_from_html(html)
            or fallback["title"]
        )
        image = (
            _extract_variant_image_from_html(html, sku_id)
            or _extract_tracking_field_from_html(html, "pdt_photo")
            or _extract_meta_content(html, "og:image")
            or _extract_meta_content(html, "twitter:image")
            or fallback["image"]
        )
        result = {
            "title": title,
            "image": image,
            "color": None,
            "variant_title": None,
        }
        _daraz_cache[cache_key] = result
        return result
    except Exception as e:
        print(f"[campaigns] Daraz lookup failed for {seller_sku or product_url}: {e}")
        _daraz_cache[cache_key] = fallback
        return fallback


def _campaign_paths(campaign_id: str) -> tuple[Path, Path]:
    safe = re.sub(r"[^a-zA-Z0-9_\-]", "", campaign_id)
    if not safe:
        abort(400, "Invalid campaign id")
    return BASE_DIR / f"{safe}.xlsx", BASE_DIR / f"{safe}.meta.json"


def _load_meta(campaign_id: str) -> dict[str, Any]:
    _, meta_path = _campaign_paths(campaign_id)
    if not meta_path.exists():
        abort(404, "Campaign not found")
    with open(meta_path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def _save_meta(campaign_id: str, meta: dict[str, Any]) -> None:
    _, meta_path = _campaign_paths(campaign_id)
    with open(meta_path, "w", encoding="utf-8") as handle:
        json.dump(meta, handle, ensure_ascii=False, indent=2)


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _to_int(value: Any) -> int | None:
    numeric = _to_float(value)
    return int(numeric) if numeric is not None else None


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _find_column_index(header_map: dict[str, int], key: str, fallback: int | None = None) -> int | None:
    for alias in HEADER_ALIASES.get(key, set()):
        if alias in header_map:
            return header_map[alias]
    return fallback


def _safe_cell(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _detect_template_columns(worksheet) -> tuple[int, dict[str, int | None]]:
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True, max_row=HEADER_SCAN_LIMIT)):
        normalized_cells = [_normalize_header(cell) for cell in row]
        header_map = {cell: idx for idx, cell in enumerate(normalized_cells) if cell}
        seller_idx = _find_column_index(header_map, "seller_sku")
        campaign_idx = _find_column_index(header_map, "campaign_price")
        if seller_idx is not None and campaign_idx is not None:
            return row_index, {
                "seller_sku": seller_idx,
                "sku_id": _find_column_index(header_map, "sku_id"),
                "shop_sku": _find_column_index(header_map, "shop_sku"),
                "product_name": _find_column_index(header_map, "product_name"),
                "product_url": _find_column_index(header_map, "product_url"),
                "sales_price": _find_column_index(header_map, "sales_price"),
                "campaign_price": campaign_idx,
                "recommended_price": _find_column_index(header_map, "recommended_price"),
                "stock": _find_column_index(header_map, "stock"),
                "is_hero": _find_column_index(header_map, "is_hero"),
                "category": _find_column_index(header_map, "category"),
            }

    return 0, dict(LEGACY_COLUMN_MAP)


def _read_template_rows(xlsx_path: Path) -> tuple[list[dict[str, Any]], dict[str, int | None]]:
    workbook = load_workbook(xlsx_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row_index, column_map = _detect_template_columns(worksheet)
    rows: list[dict[str, Any]] = []

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if row_index <= header_row_index:
            continue
        seller_sku_value = _safe_cell(row, column_map.get("seller_sku"))
        if not row or seller_sku_value in (None, ""):
            continue

        seller_sku = str(seller_sku_value).strip()
        sales_price = _to_float(_safe_cell(row, column_map.get("sales_price")))
        campaign_price = _to_float(_safe_cell(row, column_map.get("campaign_price")))
        rec_raw = _safe_cell(row, column_map.get("recommended_price"))
        rec_max = parse_recommended_max(rec_raw)
        stock_value = _safe_cell(row, column_map.get("stock"))
        is_hero_value = _safe_cell(row, column_map.get("is_hero"))

        rows.append(
            {
                "row_index": row_index,
                "seller_sku": seller_sku,
                "sku_id": str(_safe_cell(row, column_map.get("sku_id")) or ""),
                "shop_sku": str(_safe_cell(row, column_map.get("shop_sku")) or ""),
                "daraz_name": _safe_cell(row, column_map.get("product_name")) or "",
                "product_url": _safe_cell(row, column_map.get("product_url")) or "",
                "sales_price": sales_price,
                "campaign_price": campaign_price,
                "campaign_price_original": campaign_price,
                "recommended_max": rec_max,
                "recommended_raw": str(rec_raw or ""),
                "stock": _to_int(stock_value),
                "stock_original": _to_int(stock_value),
                "is_hero": (str(is_hero_value or "N").upper() == "Y"),
                "category": _safe_cell(row, column_map.get("category")) or "",
            }
        )

    return rows, column_map


campaigns_bp = Blueprint(
    "campaigns",
    __name__,
    template_folder="templates",
    static_folder="static",
    url_prefix="/campaigns",
)


@campaigns_bp.route("/")
def list_campaigns():
    items = []
    for meta_file in sorted(BASE_DIR.glob("*.meta.json"), key=os.path.getmtime, reverse=True):
        try:
            with open(meta_file, "r", encoding="utf-8") as handle:
                meta = json.load(handle)
            items.append(
                {
                    "id": meta_file.stem.replace(".meta", ""),
                    "name": meta.get("name", "Untitled"),
                    "uploaded_at": meta.get("uploaded_at", ""),
                    "original_filename": meta.get("original_filename", ""),
                    "row_count": len(meta.get("rows", [])),
                    "excluded_count": sum(1 for row in meta.get("rows", []) if row.get("excluded")),
                    "edited_count": sum(
                        1
                        for row in meta.get("rows", [])
                        if (
                            row.get("campaign_price") != row.get("campaign_price_original")
                            or row.get("stock") != row.get("stock_original")
                            or row.get("excluded")
                        )
                    ),
                }
            )
        except Exception as e:
            print(f"[campaigns] could not load {meta_file}: {e}")

    return render_template("campaigns.html", view="list", campaigns=items)


@campaigns_bp.route("/upload", methods=["POST"])
def upload_campaign():
    uploaded_file = request.files.get("file")
    name = (request.form.get("name") or "").strip() or "Untitled Campaign"
    if not uploaded_file or not uploaded_file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Please upload a .xlsx Daraz campaign template"}), 400

    campaign_id = uuid.uuid4().hex[:12]
    xlsx_path, _ = _campaign_paths(campaign_id)
    uploaded_file.save(xlsx_path)

    try:
        rows, column_map = _read_template_rows(xlsx_path)
    except Exception as e:
        xlsx_path.unlink(missing_ok=True)
        return jsonify({"error": f"Could not parse template: {e}"}), 400

    state_rows = []
    for row in rows:
        state_rows.append(
            {
                **row,
                "min_price": 0,
                "campaign_price": row["campaign_price"],
                "enrichment": None,
                "excluded": False,
            }
        )

    meta = {
        "name": name,
        "original_filename": uploaded_file.filename,
        "uploaded_at": datetime.now().isoformat(timespec="seconds"),
        "enrichment_version": ENRICHMENT_VERSION,
        "column_map": column_map,
        "rows": state_rows,
    }
    _save_meta(campaign_id, meta)
    return jsonify({"id": campaign_id, "redirect": f"/campaigns/{campaign_id}"})


@campaigns_bp.route("/<campaign_id>")
def edit_campaign(campaign_id: str):
    meta = _ensure_campaign_enrichment_version(campaign_id, _load_meta(campaign_id))
    return render_template("campaigns.html", view="edit", campaign_id=campaign_id, meta=meta)


@campaigns_bp.route("/<campaign_id>/data")
def campaign_data(campaign_id: str):
    meta = _ensure_campaign_enrichment_version(campaign_id, _load_meta(campaign_id))
    return jsonify(meta)


@campaigns_bp.route("/<campaign_id>/enrich", methods=["POST"])
def enrich_campaign(campaign_id: str):
    meta = _ensure_campaign_enrichment_version(campaign_id, _load_meta(campaign_id))
    body = request.get_json(silent=True) or {}
    requested = body.get("seller_skus")
    batch_size = int(body.get("batch_size", 10))

    enriched_now: dict[str, dict[str, Any]] = {}
    to_process: list[str] = []

    if requested:
        to_process = [sku for sku in requested if sku]
    else:
        for row in meta["rows"]:
            if row.get("enrichment") is None:
                to_process.append(row["seller_sku"])
                if len(to_process) >= batch_size:
                    break

    seen = set()
    to_process = [sku for sku in to_process if not (sku in seen or seen.add(sku))]

    for sku in to_process:
        row = next((entry for entry in meta["rows"] if entry["seller_sku"] == sku), None)
        enriched_now[sku] = enrich_from_daraz_row(row or {"seller_sku": sku})

    for row in meta["rows"]:
        if row["seller_sku"] in enriched_now:
            row["enrichment"] = enriched_now[row["seller_sku"]]

    _save_meta(campaign_id, meta)
    remaining = sum(1 for row in meta["rows"] if row.get("enrichment") is None)
    return jsonify({"enriched": enriched_now, "remaining": remaining})


@campaigns_bp.route("/<campaign_id>/update", methods=["POST"])
def update_prices(campaign_id: str):
    meta = _load_meta(campaign_id)
    body = request.get_json(silent=True) or {}
    updates = {u["seller_sku"]: u for u in body.get("updates", []) if u.get("seller_sku")}
    if not updates:
        return jsonify({"updated": 0})

    changed = 0
    for row in meta["rows"]:
        update = updates.get(row["seller_sku"])
        if not update:
            continue
        if "campaign_price" in update:
            value = _to_float(update["campaign_price"])
            if value is not None:
                row["campaign_price"] = value
                changed += 1
        if "min_price" in update:
            value = _to_float(update["min_price"])
            if value is not None:
                row["min_price"] = value
        if "stock" in update:
            value = _to_int(update["stock"])
            if value is not None:
                row["stock"] = max(0, value)
        if "excluded" in update:
            row["excluded"] = bool(update["excluded"])

    _save_meta(campaign_id, meta)
    return jsonify({"updated": changed})


@campaigns_bp.route("/<campaign_id>/bulk", methods=["POST"])
def bulk_apply(campaign_id: str):
    meta = _load_meta(campaign_id)
    body = request.get_json(silent=True) or {}
    action = body.get("action")
    params = body.get("params") or {}
    target = set(body.get("seller_skus") or [])

    def applies(row):
        return not target or row["seller_sku"] in target

    changed = 0
    for row in meta["rows"]:
        if not applies(row):
            continue
        new_price = row["campaign_price"]
        if action == "set_to_recommended_max" and row.get("recommended_max"):
            new_price = row["recommended_max"]
        elif action == "percent_off_sales":
            pct = float(params.get("pct", 0))
            if row.get("sales_price"):
                new_price = round(row["sales_price"] * (1 - pct / 100.0), 2)
        elif action == "flat_amount_off_sales":
            amount = float(params.get("amount", 0))
            if row.get("sales_price"):
                new_price = round(row["sales_price"] - amount, 2)
        elif action == "set_min_to_percent_of_sales":
            pct = float(params.get("pct", 0))
            if row.get("sales_price"):
                row["min_price"] = round(row["sales_price"] * pct / 100.0, 2)
                changed += 1
            continue
        elif action == "set_stock_absolute":
            value = max(0, int(float(params.get("value", 0))))
            row["stock"] = value
            changed += 1
            continue
        elif action == "increase_stock_by":
            value = max(0, int(float(params.get("value", 0))))
            row["stock"] = max(0, int(row.get("stock") or 0) + value)
            changed += 1
            continue
        elif action == "decrease_stock_by":
            value = max(0, int(float(params.get("value", 0))))
            row["stock"] = max(0, int(row.get("stock") or 0) - value)
            changed += 1
            continue
        elif action == "exclude_products":
            row["excluded"] = True
            changed += 1
            continue
        elif action == "restore_products":
            row["excluded"] = False
            changed += 1
            continue
        else:
            continue

        if row.get("recommended_max") and new_price > row["recommended_max"]:
            new_price = row["recommended_max"]
        if row.get("min_price") and new_price < row["min_price"]:
            new_price = row["min_price"]
        row["campaign_price"] = new_price
        changed += 1

    _save_meta(campaign_id, meta)
    return jsonify({"updated": changed})


@campaigns_bp.route("/<campaign_id>/download")
def download_campaign(campaign_id: str):
    import shutil

    meta = _load_meta(campaign_id)
    xlsx_path, _ = _campaign_paths(campaign_id)
    if not xlsx_path.exists():
        abort(404, "Source template missing")

    column_map = meta.get("column_map") or dict(LEGACY_COLUMN_MAP)
    rows_by_excel_row = {row["row_index"] + 1: row for row in meta["rows"]}

    out_path = BASE_DIR / f"{campaign_id}.download.xlsx"
    shutil.copy(xlsx_path, out_path)

    with zipfile.ZipFile(out_path) as zin:
        sheet_xml = zin.read("xl/worksheets/sheet1.xml")
        other_files = {name: zin.read(name) for name in zin.namelist() if name != "xl/worksheets/sheet1.xml"}

    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ET.register_namespace("", ns["x"])
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find("x:sheetData", ns)
    if sheet_data is None:
        abort(500, "Invalid source template: missing sheet data")

    def format_price(price):
        if price is None:
            return ""
        if isinstance(price, (int, float)):
            if float(price).is_integer():
                return str(int(price))
            return f"{price:.2f}".rstrip("0").rstrip(".")
        return str(price)

    def set_cell_value(cell, value, inline_string=False):
        for child in list(cell):
            cell.remove(child)
        if inline_string:
            cell.attrib["t"] = "inlineStr"
            is_node = ET.SubElement(cell, f"{{{ns['x']}}}is")
            text_node = ET.SubElement(is_node, f"{{{ns['x']}}}t")
            text_node.text = value
        else:
            cell.attrib.pop("t", None)
            value_node = ET.SubElement(cell, f"{{{ns['x']}}}v")
            value_node.text = value

    def excel_col_letter(index: int | None) -> str | None:
        if index is None or index < 0:
            return None
        number = index + 1
        letters = ""
        while number:
            number, remainder = divmod(number - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    campaign_col = excel_col_letter(column_map.get("campaign_price"))
    stock_col = excel_col_letter(column_map.get("stock"))

    for xml_row in list(sheet_data):
        row_num = int(xml_row.attrib.get("r", "0"))
        row_meta = rows_by_excel_row.get(row_num)
        if not row_meta:
            continue
        if row_meta.get("excluded"):
            sheet_data.remove(xml_row)
            continue

        for cell in xml_row.findall("x:c", ns):
            ref = cell.attrib.get("r", "")
            if campaign_col and ref.startswith(f"{campaign_col}{row_num}"):
                set_cell_value(cell, format_price(row_meta.get("campaign_price")), inline_string=True)
            elif stock_col and ref.startswith(f"{stock_col}{row_num}"):
                set_cell_value(cell, str(max(0, int(row_meta.get("stock") or 0))), inline_string=False)

    sheet_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in other_files.items():
            zout.writestr(name, data)
        zout.writestr("xl/worksheets/sheet1.xml", sheet_xml)

    download_name = f"{meta.get('name', 'campaign').replace(' ', '_')}_filled.xlsx"
    return send_file(
        out_path,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@campaigns_bp.route("/<campaign_id>", methods=["DELETE"])
def delete_campaign(campaign_id: str):
    xlsx_path, meta_path = _campaign_paths(campaign_id)
    xlsx_path.unlink(missing_ok=True)
    meta_path.unlink(missing_ok=True)
    (BASE_DIR / f"{campaign_id}.download.xlsx").unlink(missing_ok=True)
    return jsonify({"deleted": True})
